import os
import re
from io import StringIO

# import numpy as np
import openpyxl
from openpyxl import load_workbook

from copy import copy
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, process_pdf


def read_pdf(pdf):
    # resource manager
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    laparams = LAParams()
    # device
    device = TextConverter(rsrcmgr, retstr, laparams=laparams)
    process_pdf(rsrcmgr, device, pdf)
    device.close()
    content = retstr.getvalue()
    retstr.close()
    # 获取所有行
    lines = str(content).split("\n")
    new_name = 'testcase.txt'
    # if os.path.exists(new_name): os.remove(new_name)
    with open("%s" % (new_name), 'w', encoding="utf-8") as f:

        for line in lines:
            """移除版权行"""
            pattern = re.compile(r'\d+.*copyright.*\.', re.I)  # 注意用4个\\\\来替换\
            if pattern.match(line):
                lines.remove(line)
                continue
            """移除页码行"""
            if re.compile(r'^\d+$').match(line):
                lines.remove(line)
                continue
            """将每条用例结尾处加上感叹号方面"""
            if re.compile(r'TC\S+\d+\s').match(line) or re.compile(r'chapter', re.I).match(line):
                f.write('!!!\n')

            f.write(line + '\n')
    return lines


def find_revised(lines):
    flag_add = 0
    added = []
    flag_revise = 0
    revised = []
    flag_remove = 0
    removed = []
    for line in lines:

        if re.search('Added:', line) is not None or flag_add == 1:
            # 以逗号为分割符分割字符串，再将分割好的list元素去除前后空格和多余的字符后添加到要输出的数组中
            added.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            added = [x for x in added if x.strip() != '' or len(x.strip()) != 0]
            flag_add = 1 if line.strip().endswith(',') else 0
            continue

        if re.search('Revised:', line) is not None or flag_revise == 1:
            revised.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            revised = [x for x in revised if x.strip() != '' or len(x.strip()) != 0]
            flag_revise = 1 if line.strip().endswith(',') else 0
            continue

        if re.search('Removed:', line) is not None or flag_remove == 1:
            removed.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            removed = [x for x in removed if x.strip() != '' or len(x.strip()) != 0]
            flag_remove = 1 if line.strip().endswith(',') else 0
            continue

    return {'added': added, 'revised': revised, 'removed': removed}

    return {'added': added, 'revised': revised, 'removed': removed}


def read_tc_title(excel_path=None, old_sheet=None, sheet_name=None, revised_json=None, testcase_column_name="用例编号"):
    global f, excel_obj, writer
    txt_name = 'testcase.txt'
    revised_tc = []

    try:
        new_sheet_name = sheet_name + "（更新中）"

        # 需要用pandas去进行用例部分内容的替换，首先需要获取到sheet的dataframe
        excel_obj = pd.ExcelFile(excel_path, engine='openpyxl')
        df = pd.read_excel(excel_path, sheet_name=old_sheet)

        # 从解析出来的txt中获取到所有行
        f = open(txt_name, "r", encoding="UTF-8", errors="ignore")
        lines = f.readlines()

        # 更新增加的用例
        for tc in revised_json["added"]:
            result = analyze_test_case(lines, tc)
            pattern = re.match(r"(\D+)(\d+)", tc)

            # 查找是否有与新增的用例的前缀相同的用例
            flag = df["用例编号"].str.contains(pattern.group(1)+"\d+", regex=True, na=False)
            process = df[flag]
            num = df.index[flag]

            # 如果有，则继续查找是否有用例后缀数字比更新的用例的后缀数字大，有则获取到该用例的index；否则，就在为该模块最后一条用例的index+1
            if len(num) > 0:
                insert_id = num[-1] + 1
                row_tc = process.index[tc < process["用例编号"]]
                if len(row_tc) > 0:
                    insert_id = row_tc[0]
            # 如果没有，则在第一行进行插入
            else:
                insert_id = 1

            # 通过上述获取的插入index将原dataframe分为两份，并使用新增的用例创建dataframe，将三份dataframe合并
            # 创建dataframe时，如果时直接传入标称属性为value的字典需要写入index，也就是说，需要在创建DataFrame对象时设定index。当直接传入dict创建时，会报错Cannot mask with non-boolean array containing NA / NaN values
            df_add = pd.DataFrame({"用例编号": [tc], "用例标题": [result["title"]], "适用设备": [result["applies"]], "英文步骤": [result["content"]], "用例更新点(Revised)": [sheet_name + "新增"]})
            df1 = df.iloc[:insert_id, :]
            df2 = df.iloc[insert_id:, :]
            df = pd.concat([df1, df_add, df2], ignore_index=True, axis=0)

        # 更新更新的用例
        # f = open(txt_name, "r", encoding="gb18030", errors="ignore")  # 会导致出现中文乱码
        for tc in revised_json['revised']:
            result = analyze_test_case(lines, tc)

            # 使用pandas获取单元格行列，并替换dataframe的值
            row_tc = df.index[df["目录层级"].str.contains(tc)].tolist()  # this will only contain 2,4,6 rows
            if len(row_tc) > 0:
                df.at[row_tc, "用例标题"] = result.get('title')
                df.at[row_tc, "适用设备"] = result.get('applies')
                df.at[row_tc, "英文步骤"] = result.get('content')
                df.at[row_tc, "用例更新点(Revised)"] = sheet_name + "更新"

        # 更新移除的用例
        for tc in revised_json['removed']:
            """使用pandas获取单元格行列，并替换dataframe的值"""
            row_tc = df.index[df["用例编号"] == tc].tolist()  # this will only contain 2,4,6 rows
            if len(row_tc) > 0:
                df.at[row_tc, "用例更新点(Revised)"] = sheet_name + "移除"

        testcase_list = list()
        testcase_list.extend(revised_json["added"])
        testcase_list.append(revised_json["revised"])
        testcase_list.append(revised_json["removed"])
        # testcase_list.append(revised_json["removed"])
        df.style.applymap(case_highlihgt, ["用例编号"])
        # print(df)

        # 在不覆盖原excel的情况下，将dataframe追加写入到新的sheet中
        book = openpyxl.load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine="openpyxl")
        writer.book = book
        if new_sheet_name in writer.book.sheetnames:
            writer.book.remove(writer.book[new_sheet_name])
        df.to_excel(excel_writer=writer, sheet_name=new_sheet_name, index=None)
        writer.save()
        writer.close()
        book.close()

    finally:
        f.close()
        excel_obj.close()
    return revised_tc


def case_highlihgt(row, case_list):
    print(row, case_list)
    return ["background-color:yellow" if i in case_list else "background-color:white" for i in row]


def analyze_test_case(lines, testcase_name) -> dict:
    result = dict()

    flag_title = flag_tc = flag_applies = flag_empty = 0
    title = ''
    case_content = ''
    applies = ''

    it = iter(lines)
    for line in it:
        if line.find(testcase_name + ' ') > -1 or flag_title == 1 or flag_applies == 1 or flag_tc == 1:

            if line.find(testcase_name + ' ') > -1:
                flag_title = 1
                title = title + line.replace(testcase_name + ' ', '')
                continue
            if flag_title == 1:
                title = title + line
                if line.strip() == '':
                    flag_applies = 1
                    flag_title = 0
                    continue
                # it.__next__()
            # elif line.find('Applies to') > -1:
            if flag_applies == 1:
                if line.strip() == '':
                    flag_applies = 0
                    flag_tc = 1
                    continue
                else:
                    applies = applies + line
                    continue
            if flag_tc == 1:
                if line.strip() == '':
                    flag_empty = flag_empty + 1
                    continue
                if line.startswith('!!!') or flag_empty > 1:
                    flag_tc = 0
                    break
                else:
                    case_content = case_content + line
                    flag_empty = 0

    """将部分不是常用字符的内容删除"""
    title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
    case_content = ILLEGAL_CHARACTERS_RE.sub(r'', case_content)
    applies = ILLEGAL_CHARACTERS_RE.sub(r'', applies)

    return {'tc': testcase_name, 'title': title, 'applies': applies, 'content': case_content}


if __name__ == '__main__':
    pdf_path = "/Users/zaochuan/Downloads/HomeKit Certification Test Cases R11.2.pdf"
    excel_path = "./HomeKit用例.xlsx"
    old_sheet_name = "R11.1"
    new_sheet_name = "R11.2"
    with open(pdf_path, "rb") as my_pdf:
        read_tc_title(excel_path, old_sheet_name, new_sheet_name, find_revised(read_pdf(my_pdf)))
