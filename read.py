import os
import re
from io import StringIO

import openpyxl
from openpyxl import load_workbook
import xlsxwriter

from copy import copy
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, process_pdf


# def replace_excel(folder_path, file_name):
#     """
#     excel  .xls 后缀 改成 .xlsx 后缀
#     folder_path 文件夹路径
#     file_name 文件名字 带后缀 比如 aa.xls
#     """
#     name, suffix = file_name.split('.')
#     excel_file_path = os.path.join(folder_path, file_name)
#     import win32com.client
#     excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
#     wb = excel.Workbooks.Open(excel_file_path)
#     suffix = f".{suffix}x"
#     new_file_name = f"{name}{suffix}"
#     new_excel_file_path = os.sep.join([folder_path, new_file_name])
#     wb.SaveAs(new_excel_file_path, FileFormat=51)
#     wb.Close()
#     excel.Application.Quit()
#     os.remove(excel_file_path)


def read_pdf(pdf):
    """
    读取pdf的内容解析后输出
    :param pdf: pdf的文件路径
    :return: pdf每一行字符串组成的list数组
    """

    # resource manager
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    laparams = LAParams()
    # device
    device = TextConverter(rsrcmgr, retstr, laparams=laparams)
    process_pdf(rsrcmgr, device, pdf)
    device.close()
    content = retstr.getvalue()
    retstr.close()
    # 获取所有行
    lines = str(content).split("\n")
    new_name = 'testcase.txt'
    # if os.path.exists(new_name): os.remove(new_name)
    with open("%s" % (new_name), 'w', encoding="utf-8") as f:

        for line in lines:
            """移除版权行"""
            pattern = re.compile(r'\d+.*copyright.*\.', re.I)  # 注意用4个\\\\来替换\
            if pattern.match(line):
                lines.remove(line)
                continue
            """移除页码行"""
            if re.compile(r'^\d+$').match(line):
                lines.remove(line)
                continue
            """在每个2级目录前加入~~~字符以便解析"""
            if re.compile(r'\d+\.\d+\s.*').match(line.strip()):
                f.write('~~~\n')
            """将每条用例结尾处加上感叹号以便后续解析"""
            if re.compile(r'TC\S+\d+\s').match(line) or re.compile(r'chapter', re.I).match(line):
                f.write('!!!\n')

            f.write(line + '\n')
    return lines


def find_revised(lines) -> dict:
    """
    解析文档中标注的更新的用例编号，包含新增、删除、更新的用例
    :param lines: pdf中解析出的每一行内容组成的list
    :return: 新增、删除、更新的用例编号
    """
    flag_add, flag_revise, flag_remove = False, False, False
    added, revised, removed = list(), list(), list()

    def parse(flag: bool, line: str, return_list: list):
        """
        如果存在要查找的字符，则返回解析的list以及flag
        :param flag:
        :param line:
        :param return_list:
        :param find_str:
        :return: 如果存在要查找的
        """
        # 以逗号为分割符分割字符串，再将分割好的list元素去除前后空格和多余的字符后添加到要输出的数组中
        return_list.extend(
            [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
        return_list = [x for x in return_list if x.strip() != '' or len(x.strip()) != 0]
        flag = True if line.strip().endswith(',') else False
        return return_list, flag

    for line in lines:
        if re.search("Added:", line) is not None or flag_add is True:
            added, flag_add = parse(flag_add, line, added)
            continue

        if re.search("Revised:", line) is not None or flag_revise is True:
            revised, flag_revise = parse(flag_revise, line, revised)
            continue

        if re.search("Removed:", line) is not None or flag_remove is True:
            removed, flag_remove = parse(flag_remove, line, removed)
            continue

    return {'added': added, 'revised': revised, 'removed': removed}


def read_tc_title(excel_path=None, old_sheet=None, sheet_name=None, revised_json=None):
    """
    解析txt文件并且将解析后的内容输出到excel文件中
    :param excel_path: 下载的线上用例的路径，包含文件名及文件后缀
    :param old_sheet: 需要被替换的原先的sheet名称
    :param sheet_name: 本次更新的用例版本
    :param revised_json: 获取到的恶更新后的用例编号dict
    :return:
    """
    global f, writer, book
    txt_name = 'testcase.txt' # 解析pdf时生成的txt文件
    revised_tc = list()

    try:
        # 如果线上用例下载下来的文件后缀时xls，则默认根据原有的xls文件生成一份新的xlsx文件
        df = pd.read_excel(excel_path, old_sheet, header=None).iloc[1:]  # 线上下载的用例会存在一行提示，此处不保存该行提示
        if excel_path.endswith("xls"):
            excel_path = excel_path + "x"
            df.to_excel(excel_path, old_sheet, header=False, index=False, engine='xlsxwriter')

        # 最后更新后的sheet名称
        new_sheet_name = sheet_name + "（更新中）"

        # 需要用pandas去进行用例部分内容的替换，首先需要获取到sheet的dataframe
        df = pd.read_excel(excel_path, sheet_name=old_sheet)

        # 从解析出来的txt中获取到所有行
        f = open(txt_name, "r", encoding="UTF-8", errors="ignore")
        lines = f.readlines()

        # 首先先获取到目录层级，解析只限于到二级目录与用例编号
        catalog = dict()
        flag, flag_cata = 0, 0
        cata = ""
        for line in lines:
            line = line.strip()
            if re.search('~~~', line) is not None:
                flag_cata, flag = 1, 1
                continue
            if flag == 1:
                if flag_cata == 1:
                    cata = line.strip()
                    catalog[cata] = list()
                    flag_cata = 0
                if re.search("!!!", line) is not None:
                    flag = 0
                    continue
                if re.search(r'TC\S+\d+:\s', line) is not None:
                    catalog[cata].append(re.match(r'(TC\S+\d+)(.*)', line).group(1))

        # 将所有二级目录下list长度为0的元素剔除
        for k in list(catalog.keys()):
            if len(catalog[k]) == 0:
                del catalog[k]

        # 更新增加的用例
        for tc in revised_json["added"]:
            result = analyze_test_case(lines, tc)
            pattern = re.match(r"(\D+)(\d+)", tc)

            insert_id = 1

            # 创建dataframe时，如果时直接传入标称属性为value的字典需要写入index，也就是说，需要在创建DataFrame对象时设定index。当直接传入dict创建时，会报错Cannot mask with non-boolean array containing NA / NaN values
            df_add = pd.DataFrame({
                "目录层级": [tc],
                "标题*": [result["title"]],
                "前置条件": [result["applies"]],
                "步骤描述": [result["content"]],
                "用例标签": [sheet_name + "新增"],
                "预期结果": [result["content"]]
            })
            df1 = df.iloc[:insert_id, :]
            df2 = df.iloc[insert_id:, :]
            df = pd.concat([df1, df_add, df2], ignore_index=True, axis=0)

        # 更新更新的用例
        for tc in revised_json['revised']:
            result = analyze_test_case(lines, tc)

            # 使用pandas获取单元格行，并替换dataframe的值
            row_tc = df.index[df["目录层级"].str.contains(tc, na=True)].tolist()  # this will only contain 2,4,6 rows

            if len(row_tc) > 0:
                df.at[row_tc, "标题*"] = result.get('title')
                df.at[row_tc, "前置条件"] = result.get('applies')
                df.at[row_tc, "步骤描述"] = result["content"]
                df.at[row_tc, "预期结果"] = result.get('content')
                df.at[row_tc, "用例标签"] = sheet_name + "更新"

        # 更新移除的用例
        for tc in revised_json['removed']:
            """使用pandas获取单元格行列，并替换dataframe的值"""
            row_tc = df.index[df["目录层级"].str.contains(tc, na=True)].tolist()  # this will only contain 2,4,6 rows
            if len(row_tc) > 0:
                df.drop(index=row_tc)

        # 重新替换所有的用例的目录层级，避免因为大版本更新导致目录层级大调整所带来的人工更新成本
        for k, v in catalog.items():
            for item in v:
                # df.loc[df["目录层级"].str.contains(item, regex=True, na=True), "|{}|{}".format(k, item)]
                for row in df.index[df["目录层级"].str.contains(item, regex=True, na=True)]:
                    quest = str(df.at[row, "目录层级"]).strip()
                    template_exist = re.compile(r"^R\d+\.\d+(\|.*\|)+\d+.*\|.*\d$")
                    template_added = re.compile(r"^TC.*\d$")  # 新增的用例的目录层级只有用例编号
                    if re.search(template_added, quest):
                        df.at[row, "目录层级"] = re.sub(template_added, r'{}|{}'.format(k, item), quest)
                        continue
                    df.at[row, "目录层级"] = re.sub(template_exist, r'{}\g<1>{}|{}'.format(sheet_name, k, item), quest)

        # 在不覆盖原excel的情况下，将dataframe追加写入到新的sheet中
        book = openpyxl.load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine="openpyxl")
        writer.book = book
        if new_sheet_name in writer.book.sheetnames:
            writer.book.remove(writer.book[new_sheet_name])
        testcase_list = list()
        testcase_list.extend(revised_json["added"])
        testcase_list.extend(revised_json["revised"])
        testcase_list.extend(revised_json["removed"])

        def style_apply(value, **kwargs):
            """

            :param series: 传过来的数据是DataFrame中的一列   类型为pd.Series
            :param colors: 内容是字典  其中key 为标题名   value 为颜色
            :return:
            """
            back_ground = 'background-color: white'
            print("{}, {}".format(type(value), back_ground))
            for item in kwargs["cases"]:
                if value.find(item) > 0:
                    back_ground = 'background-color: yellow'

            print("{}, {}".format(value, back_ground))
            return back_ground

        # axis =0 ，按列设置样式
        df.style.applymap(style_apply, ["目录层级"], colors="yellow", cases=testcase_list, df=df).to_excel(excel_writer=writer, sheet_name=new_sheet_name, index=None)
        writer.book.active = writer.book[new_sheet_name]
        writer.save()
        writer.close()
        book.close()

    finally:
        f.close()
    return revised_tc


def case_highlihgt(row, case_list):
    return ["background-color:yellow" if i in case_list else "background-color:white" for i in row]


def analyze_test_case(lines, testcase_name) -> dict:
    """
    将传入的包含每一行字符串数组匹配用例编号进行解析
    :param lines: 字符串数组，每一个元素都是文件中的一行内容
    :param testcase_name: 用例编号
    :return: 用例的编号、标题、适用设备、内容
    """
    flag_title = flag_tc = flag_applies = flag_empty = 0
    title = ''
    case_content = ''
    applies = ''

    it = iter(lines)
    for line in it:
        if line.find(testcase_name + ' ') > -1 or flag_title == 1 or flag_applies == 1 or flag_tc == 1:

            if line.find(testcase_name + ' ') > -1:
                flag_title = 1
                title = title + line.replace(testcase_name + ' ', '')
                continue
            if flag_title == 1:
                title = title + line
                if line.strip() == '':
                    flag_applies = 1
                    flag_title = 0
                    continue
                # it.__next__()
            # elif line.find('Applies to') > -1:
            if flag_applies == 1:
                if line.strip() == '':
                    flag_applies = 0
                    flag_tc = 1
                    continue
                else:
                    applies = applies + line
                    continue
            if flag_tc == 1:
                if line.strip() == '':
                    flag_empty = flag_empty + 1
                    continue
                if line.startswith('!!!') or flag_empty > 1:
                    flag_tc = 0
                    break
                else:
                    case_content = case_content + line
                    flag_empty = 0

    """将部分不是常用字符的内容删除"""
    title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
    case_content = ILLEGAL_CHARACTERS_RE.sub(r'', case_content)
    applies = ILLEGAL_CHARACTERS_RE.sub(r'', applies)

    return {'tc': testcase_name, 'title': title, 'applies': applies, 'content': case_content}


if __name__ == '__main__':
    pdf_path = "/Users/zaochuan/Downloads/HomeKit Certification Test Cases R11.2.pdf"
    excel_path = "/Users/zaochuan/Downloads/线上用例库R11.1.xls"
    old_sheet_name = "用例"
    new_sheet_name = "R11.2"
    with open(pdf_path, "rb") as my_pdf:
        read_tc_title(excel_path, old_sheet_name, new_sheet_name, find_revised(read_pdf(my_pdf)))
