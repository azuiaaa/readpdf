import os
import re
from io import StringIO

import numpy as np
import openpyxl
import xlrd
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, process_pdf


def read_pdf(pdf):
    # resource manager
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    laparams = LAParams()
    # device
    device = TextConverter(rsrcmgr, retstr, laparams=laparams)
    process_pdf(rsrcmgr, device, pdf)
    device.close()
    content = retstr.getvalue()
    retstr.close()
    # 获取所有行
    lines = str(content).split("\n")
    new_name = 'testcase.txt'
    if os.path.exists(new_name): os.remove(new_name)
    with open("%s" % (new_name), 'a', encoding="utf-8") as f:
        # new_lines = set(lines)
        # content = re.sub('\d+.*copyright.*\.[\r|\r\n].*\n\d*', '', content, count=0, flags=re.I)
        # pattern = re.compile('\d+.*copyright.*\.[\r|\r\n].*\n\d*')

        # print(pattern.search(content))
        # f.write(content)
        for line in lines:
            # line.strip().replace('\d+.*copyright.*\.', '')
            # re.sub('', )
            pattern = re.compile(r'\d+.*copyright.*\.', re.I)  # 注意用4个\\\\来替换\
            # line = pattern.sub('', line.strip())
            if pattern.match(line):
                # print(line)
                lines.remove(line)
                continue
            if re.compile(r'^\d+$').match(line):
                # print(line)
                lines.remove(line)
                continue
            # line = re.compile(r'^\d+$').sub('', line.strip())
            if re.compile(r'TC\S+\d+\s').match(line) or re.compile(r'chapter', re.I).match(line):
                print(line)
                f.write('!!!\n')

            f.write(line+'\n')
    return lines


def find_revised(lines):
    flag_add = 0
    added = []
    flag_revise = 0
    revised = []
    flag_remove = 0
    removed = []
    for line in lines:

        if re.search('Added:', line) is not None or flag_add == 1:
            added.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            flag_add = 1 if line.strip().endswith(',') else 0
            continue

        if re.search('Revised:', line) is not None or flag_revise == 1:
            revised.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            flag_revise = 1 if line.strip().endswith(',') else 0
            continue

        if re.search('Removed:', line) is not None or flag_remove == 1:
            removed.extend(
                [x.strip().split(' ')[2] if len(x.strip().split(' ')) > 1 else x.strip() for x in line.split(',')])
            flag_remove = 1 if line.strip().endswith(',') else 0
            continue

    return {'added': added, 'revised': revised, 'removed': removed}


def read_tc_title(excel_path=None, sheet_name=None, revised_json=None):
    global f
    txt_name = 'testcase.txt'
    # data = openpyxl.load_workbook(excel_path)  # 打开工作簿
    # sheetnames = data.get_sheet_names()
    # data.create_sheet('a', 0)
    # sheet = data.get_sheet_by_name(sheet_name)
    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    # sheet = data.active
    revised_tc = []
    try:
        f = open(txt_name, "r")
        lines = f.readlines()
        for tc in revised_json['revised']:
            flag_title = flag_tc = flag_applies = flag_empty = 0
            title = ''
            case_content = ''
            # print(tc)
            sum = 0
            it = iter(lines)
            for line in it:
                if line.find(tc + ' ') > -1 or flag_title == 1 or flag_applies == 1 or flag_tc == 1:

                    if line.find(tc + ' ') > -1:
                        # sum = sum + 1
                        # print(sum)
                        flag_title = 1
                        title = title + line.replace(tc + ' ', '')
                        continue
                    if flag_title == 1:
                        title = title + line
                        if line.strip() == '':
                            flag_applies = 1
                            flag_title = 0
                            continue
                        # it.__next__()
                    # elif line.find('Applies to') > -1:
                    if flag_applies == 1:
                        if line.strip() == '':
                            flag_applies = 0
                            flag_tc = 1
                            continue
                        else:
                            continue
                    if flag_tc == 1:
                        if line.strip() == '':
                            flag_empty = flag_empty + 1
                            continue
                        if line.startswith('!!!') or flag_empty > 1:
                            flag_tc = 0
                            break
                        else:
                            case_content = case_content + line
                            flag_empty = 0

            title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
            case_content = ILLEGAL_CHARACTERS_RE.sub(r'', case_content)
            revised_tc.append({'tc': tc, 'title': title, 'content': case_content})
            # nrows = sheet.max_row  # 获得行数
            # ncolumns = sheet.max_column  # 获得列数
            # 注意行业列下标是从1开始的
            row_tc = df.index[df["用例编号"] == tc].tolist()  # this will only contain 2,4,6 rows
            if len(row_tc) > 0:
                df.at[row_tc, "用例标题"] = title
                df.at[row_tc, "英文步骤"] = case_content

                def equal(x, color='blue'):
                    # if x == tc:
                    color = '#99ff66'  # light green  '#99ff66'
                    # else:
                    #     color = "#000000"
                    return f'background-color: {color}'

                # axis =0 ，按列设置样式
                df.style.apply(equal, axis=0)
                print(df)
                # df_consume.style.hide_index() \
                #     .hide_columns(['性别', '基金经理', '上任日期', ]) \
                #     .format(format_dict) \
                #     .apply(max_value, axis=0, subset=['2018', '2019', '2020', '2021'])

            # for row in df_tc.iterrows():
            #     row
            # sheet.cell(nrows + 1, 1).value = tc
            # sheet.cell(nrows + 1, 2).value = title
            # sheet.cell(nrows + 1, 3).value = case_content
        # data.save('Homekit用例_1.xlsx')
        df.to_excel(excel_path, sheet_name)

    finally:
        f.close()
    return revised_tc


if __name__ == '__main__':
    with open('/Users/zaochuan/Downloads/HomeKit Certification Test Cases R11.1.pdf', "rb") as my_pdf:
        read_tc_title("./HomeKit用例.xlsx", "R11.2", revised_json=find_revised(read_pdf(my_pdf)))
